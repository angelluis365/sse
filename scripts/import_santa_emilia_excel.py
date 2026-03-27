from __future__ import annotations

import argparse
import re
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from datetime import datetime

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font
from pypdf import PdfReader


DEFAULT_SOURCE_DIR = Path("/Users/angelluisabujamartinez/Downloads/neumaticos")
DEFAULT_OUTPUT_XLSX = Path("/Users/angelluisabujamartinez/Documents/New project/facturas_santa_emilia_neumaticos.xlsx")

INVOICE_RE = re.compile(r"FACTURA\s+#?(Factura_[A-Z]-\d+/\d+)")
DATE_RE = re.compile(r"Fecha:\s+(\d{2}/\d{2}/\d{4})")
PAGE_RE = re.compile(r"P[á\x80]g\.\s*(\d+)\s+de\s+(\d+)")
PRICE_LINE_RE = re.compile(
    r"(?P<unit>\d[\d\.,]*)€?\s+"
    r"(?P<qty>\d[\d\.,]*)\s+"
    r"(?:(?P<dto>\d+%)\s+)?"
    r"(?P<subtotal>\d[\d\.,]*)€?\s+"
    r"\d+%\s+"
    r"(?P<total>\d[\d\.,]*)€?"
)
SIZE_RE = re.compile(r"\b\d{3}/\d{2}(?:R|VR|HR|TR)\d{2}(?:\.\d)?|\b\d+\.\d{2}[- ]\d+\b|\bR\d{2}(?:\.\d)?\b")
REFERENCE_PREFIX_RE = re.compile(r"^([A-Z0-9.-]{4,})(CUB\b.*)$")
REFERENCE_ONLY_RE = re.compile(r"^[A-Z0-9.-]{4,}$")


@dataclass
class LineItem:
    factura: str
    fecha: str
    referencia: str
    concepto: str
    cantidad: Decimal
    precio_unitario: Decimal
    total_linea: Decimal
    archivo_pdf: str


def looks_obfuscated(text: str) -> bool:
    if "6XPLQLVWURV" in text or "&21&(372" in text:
        return True
    if any(ord(char) < 32 and char not in "\n\r\t" for char in text):
        return True
    compact = text.replace(" ", "")
    return bool(compact) and bool(re.fullmatch(r"[A-Z0-9%$&(),./:+='\-]{8,}", compact)) and any(ch in compact for ch in "&$%")


def maybe_decode_text(text: str) -> str:
    if "Suministros Santa Emilia" in text:
        return text
    if not looks_obfuscated(text):
        return text
    return "".join("\n" if char == "\n" else chr(ord(char) + 29) for char in text)


def parse_decimal(value: str) -> Decimal:
    cleaned = re.sub(r"[^\d,.\-]", "", value).replace(".", "").replace(",", ".").strip()
    return Decimal(cleaned)


def normalize_line(line: str) -> str:
    line = line.replace("\x83", "€").replace("\x7f", "€").replace("\x84", "€")
    line = re.sub(r"(?<=\d)[fgb](?=\s|$)", "€", line)
    line = line.replace("= ", " ").replace(" =", " ").replace("=", " ")
    line = line.replace(" ⁰", "").replace("º", "")
    return re.sub(r"\s+", " ", line).strip()


def split_reference(desc_lines: list[str]) -> tuple[str, str]:
    if not desc_lines:
        return "", ""
    if len(desc_lines) >= 2 and REFERENCE_ONLY_RE.fullmatch(desc_lines[0]):
        return desc_lines[0], " ".join(desc_lines[1:])
    match = REFERENCE_PREFIX_RE.match(desc_lines[0])
    if match:
        return match.group(1), " ".join([match.group(2), *desc_lines[1:]]).strip()
    concept = " ".join(desc_lines)
    return "", concept


def parse_table_row(row: list[str], factura: str, fecha: str, archivo_pdf: str) -> LineItem | None:
    decoded = [maybe_decode_text(cell or "") for cell in row]
    normalized = [normalize_line(cell) for cell in decoded]
    if not normalized or not normalized[0]:
        return None
    if "CONCEPTO" in normalized[0].upper():
        return None

    concept_raw = decoded[0]
    desc_lines = [normalize_line(line) for line in concept_raw.splitlines() if normalize_line(line)]
    referencia, concepto = split_reference(desc_lines)

    numeric_cells = [cell for cell in normalized[1:] if cell]
    if len(numeric_cells) < 5:
        return None
    if len(numeric_cells) == 5:
        unit, qty, subtotal, iva, total = numeric_cells
    else:
        unit, qty, *_middle, subtotal, iva, total = numeric_cells

    try:
        return LineItem(
            factura=factura,
            fecha=fecha,
            referencia=referencia,
            concepto=concepto,
            cantidad=parse_decimal(qty),
            precio_unitario=parse_decimal(unit),
            total_linea=parse_decimal(total),
            archivo_pdf=archivo_pdf,
        )
    except (InvalidOperation, ValueError):
        return None


def extract_items_from_text(text: str, factura: str, fecha: str, archivo_pdf: str) -> list[LineItem]:
    items: list[LineItem] = []
    current_desc: list[str] = []
    in_table = False
    for raw_line in text.splitlines():
        line = normalize_line(raw_line)
        if not line:
            continue
        if "CONCEPTO" in line and "PRECIO" in line:
            in_table = True
            current_desc = []
            continue
        if not in_table:
            continue
        if line.startswith("BASE IMPONIBLE"):
            break
        if re.match(r"^pedido de .+\d[\d\.,]*€$", line, re.IGNORECASE):
            continue
        remainder = line
        while remainder:
            match = PRICE_LINE_RE.search(remainder)
            if not match:
                current_desc.append(remainder.strip())
                break
            prefix = normalize_line(remainder[: match.start()])
            if prefix:
                current_desc.append(prefix)
            referencia, concepto = split_reference(current_desc)
            if concepto:
                items.append(
                    LineItem(
                        factura=factura,
                        fecha=fecha,
                        referencia=referencia,
                        concepto=concepto,
                        cantidad=parse_decimal(match.group("qty")),
                        precio_unitario=parse_decimal(match.group("unit")),
                        total_linea=parse_decimal(match.group("total")),
                        archivo_pdf=archivo_pdf,
                    )
                )
            current_desc = []
            remainder = normalize_line(remainder[match.end() :])
    return items


def extract_items(pdf_path: Path) -> list[LineItem]:
    reader = PdfReader(str(pdf_path))
    raw_text = "\n".join((page.extract_text() or "") for page in reader.pages)
    text = maybe_decode_text(raw_text)
    if "Suministros Santa Emilia" not in text:
        return []
    invoice_match = INVOICE_RE.search(text)
    date_match = DATE_RE.search(text)
    factura = invoice_match.group(1) if invoice_match else pdf_path.stem
    fecha = date_match.group(1) if date_match else ""
    items: list[LineItem] = []
    if "6XPLQLVWURV" in raw_text:
        items = extract_items_from_text(text, factura, fecha, pdf_path.name)
        if items:
            return items
    seen_page_numbers: set[str] = set()
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            page_text = maybe_decode_text(page.extract_text() or "")
            if "Suministros Santa Emilia" not in page_text:
                continue
            page_match = PAGE_RE.search(page_text)
            page_number = page_match.group(1) if page_match else ""
            if page_number and page_number in seen_page_numbers:
                continue
            if page_number:
                seen_page_numbers.add(page_number)
            for table in page.extract_tables() or []:
                for row in table:
                    item = parse_table_row(row, factura, fecha, pdf_path.name)
                    if item is not None:
                        items.append(item)
    if not items:
        items = extract_items_from_text(text, factura, fecha, pdf_path.name)
    return items


def autofit(ws) -> None:
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 60)


def build_workbook(items: list[LineItem], output_xlsx: Path) -> None:
    wb = Workbook()

    ws_grouped = wb.active
    ws_grouped.title = "Agrupado"
    ws_grouped.append(["Referencia agrupada", "Concepto", "Cantidad total", "Precio total", "Facturas"])

    grouped: dict[tuple[str, str], dict[str, object]] = defaultdict(lambda: {"cantidad": Decimal("0"), "total": Decimal("0"), "facturas": set()})
    for item in items:
        key_ref = item.referencia or item.concepto
        key = (key_ref, item.concepto)
        grouped[key]["cantidad"] += item.cantidad
        grouped[key]["total"] += item.total_linea
        grouped[key]["facturas"].add(item.factura)

    for (ref, concepto), data in sorted(grouped.items(), key=lambda row: (row[0][1], row[0][0])):
        ws_grouped.append(
            [
                ref,
                concepto,
                float(data["cantidad"]),
                float(data["total"]),
                ", ".join(sorted(data["facturas"])),
            ]
        )

    ws_year = wb.create_sheet("Ano y concepto")
    ws_year.append(["Ano", "Referencia agrupada", "Concepto", "Cantidad total", "Precio total", "Facturas"])
    grouped_year: dict[tuple[str, str, str], dict[str, object]] = defaultdict(
        lambda: {"cantidad": Decimal("0"), "total": Decimal("0"), "facturas": set()}
    )
    for item in items:
        year = datetime.strptime(item.fecha, "%d/%m/%Y").year if item.fecha else ""
        key_ref = item.referencia or item.concepto
        key = (str(year), key_ref, item.concepto)
        grouped_year[key]["cantidad"] += item.cantidad
        grouped_year[key]["total"] += item.total_linea
        grouped_year[key]["facturas"].add(item.factura)

    for (year, ref, concepto), data in sorted(grouped_year.items(), key=lambda row: (row[0][0], row[0][2], row[0][1])):
        ws_year.append(
            [
                year,
                ref,
                concepto,
                float(data["cantidad"]),
                float(data["total"]),
                ", ".join(sorted(data["facturas"])),
            ]
        )

    ws_detail = wb.create_sheet("Detalle")
    ws_detail.append(["Factura", "Fecha", "Referencia", "Concepto", "Cantidad", "Precio unitario", "Total linea", "Archivo PDF"])
    for item in sorted(items, key=lambda row: (row.concepto, row.factura, row.referencia)):
        ws_detail.append(
            [
                item.factura,
                item.fecha,
                item.referencia,
                item.concepto,
                float(item.cantidad),
                float(item.precio_unitario),
                float(item.total_linea),
                item.archivo_pdf,
            ]
        )

    for ws in (ws_grouped, ws_year, ws_detail):
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column in {3, 4, 5, 6, 7} and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
        ws.freeze_panes = "A2"
        autofit(ws)

    wb.save(output_xlsx)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-dir", default=str(DEFAULT_SOURCE_DIR))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT_XLSX))
    args = parser.parse_args()

    source_dir = Path(args.source_dir)
    output_xlsx = Path(args.output)
    all_items: list[LineItem] = []
    for pdf_path in sorted(source_dir.glob("*.pdf")):
        all_items.extend(extract_items(pdf_path))
    build_workbook(all_items, output_xlsx)
    print(f"PDFs procesados: {len(list(source_dir.glob('*.pdf')))}")
    print(f"Lineas importadas: {len(all_items)}")
    print(f"Excel generado: {output_xlsx}")


if __name__ == "__main__":
    main()
